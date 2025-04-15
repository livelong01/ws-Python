# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
# Tutorial: https://openpyxl.readthedocs.io/en/stable/tutorial.html

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from pathlib import Path


ROOT_DIR = Path(__file__).parent
WORKBOOK_PATH = ROOT_DIR / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active

# nomeou a planilha
sheet_name = 'Minha Planilha'
# cria uma nova planilha na posição 0 (primeira)
workbook.create_sheet(sheet_name, 0)  # workbook.create_sheet(sheet_name, 0)
# worksheet: Worksheet = workbook.active  # type: ignore
# selecionou a planilha criada
worksheet: Worksheet = workbook[sheet_name]  # type: ignore
# remover a planilha padrão criada automaticamente
workbook.remove(workbook['Sheet'])  # type: ignore
print(workbook.sheetnames)

# criando o cabeçalho
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # name      age  grade
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
    ['Ana',     12,   7.5],
    ['Marcos',  14,   6.0],
    ['Juliana', 13,   8.2],
    ['Fernanda', 15,   9.0],
    ['Carlos',  16,   7.8],
    ['Patrícia', 12,   9.5],
]


# for i, students_row in enumerate(students, start=2):
#     for j, students_column in enumerate(students_row, start=1):
#         worksheet.cell(i, j, students_column)

for student in students:
    worksheet.append(student)


workbook.save(WORKBOOK_PATH)
