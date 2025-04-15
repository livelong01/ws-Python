# openpyxl - Ler e alterar dados de uma planilha.
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
# Tutorial: https://openpyxl.readthedocs.io/en/stable/tutorial.html

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from pathlib import Path


ROOT_DIR = Path(__file__).parent
WORKBOOK_PATH = ROOT_DIR / 'workbook.xlsx'

workbook: Workbook = load_workbook(WORKBOOK_PATH)


# nomeou a planilha
sheet_name = 'Minha Planilha'

# selecionou a planilha criada
worksheet: Worksheet = workbook[sheet_name]  # type: ignore

row: tuple[Cell]

for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    for cell in row:
        print(cell.value, end='\t')
        # outra forma de alterar o valor de uma célula
        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, 15)
    print()
    workbook.save(WORKBOOK_PATH)

# Alterando o valor de uma célula
# worksheet['B3'].value = 14


workbook.save(WORKBOOK_PATH)
workbook.close()
